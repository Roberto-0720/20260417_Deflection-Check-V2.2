"""Export deflection check results to Excel and TXT files."""
import os
from datetime import datetime
from typing import List
from data.models import BeamCheckSummary

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    HEADER_FILL = PatternFill("solid", fgColor="2F5496") if HAS_OPENPYXL else None
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial") if HAS_OPENPYXL else None
    OK_FILL = PatternFill("solid", fgColor="C6EFCE") if HAS_OPENPYXL else None
    NG_FILL = PatternFill("solid", fgColor="FFC7CE") if HAS_OPENPYXL else None
    OK_FONT = Font(bold=True, color="006100", name="Arial", size=10) if HAS_OPENPYXL else None
    NG_FONT = Font(bold=True, color="9C0006", name="Arial", size=10) if HAS_OPENPYXL else None
    TITLE_FONT = Font(bold=True, size=14, name="Arial") if HAS_OPENPYXL else None
    NORMAL_FONT = Font(size=10, name="Arial") if HAS_OPENPYXL else None
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    ) if HAS_OPENPYXL else None

    def export(self, results: List[BeamCheckSummary], output_dir: str,
               allowable_ratio: float = 360, project_name: str = "",
               abs_limit_mm: float = 25.0,
               use_rel: bool = True, use_abs: bool = True) -> str:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required")
        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(output_dir, f"Deflection_Check_{ts}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Deflection Check Summary"

        # Build dynamic column list
        # Base columns always present
        base_headers = ["Beam (Group)", "Section", "Ctrl Load Combo", "Span (mm)"]
        base_widths   = [16, 22, 16, 12]
        rel_headers   = ["Deflection", "Max Defl (mm)", "Criteria", "Rel"]
        rel_widths     = [16, 14, 12, 8]
        abs_headers   = ["Abs Max (mm)", f"Abs Limit ({abs_limit_mm}mm)", "Abs"]
        abs_widths     = [14, 16, 8]
        tail_headers  = ["Ctrl Node", "Element List"]
        tail_widths    = [14, 30]

        headers = base_headers
        col_widths = base_widths
        if use_rel:
            headers   += rel_headers
            col_widths += rel_widths
        if use_abs:
            headers   += abs_headers
            col_widths += abs_widths
        headers   += tail_headers
        col_widths += tail_widths

        total_cols = len(headers)
        last_col_letter = get_column_letter(total_cols)

        criteria_str = []
        if use_rel:
            criteria_str.append(f"Deflection Limit: L/{int(allowable_ratio)}")
        if use_abs:
            criteria_str.append(f"Abs. Limit: {abs_limit_mm} mm")

        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = f"DEFLECTION CHECK SUMMARY — {project_name}" if project_name else "DEFLECTION CHECK SUMMARY"
        ws["A1"].font = self.TITLE_FONT
        ws["A2"] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10, name="Arial")
        ws["A3"] = "   |   ".join(criteria_str)
        ws["A3"].font = Font(italic=True, size=10, name="Arial")

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=5, column=ci, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[5].height = 30

        def _sort_key(x):
            n = x.group_name.strip()
            return (0, int(n)) if n.isdigit() else (1, n)

        def _is_ok(r):
            if use_rel and use_abs:
                return r.rel_is_ok and r.abs_is_ok
            if use_rel:
                return r.rel_is_ok
            return r.abs_is_ok

        row = 6
        for r in sorted(results, key=_sort_key):
            col = 1
            ws.cell(row=row, column=col, value=r.group_name).font = self.NORMAL_FONT;  col += 1
            ws.cell(row=row, column=col, value=r.section).font = self.NORMAL_FONT;      col += 1
            ws.cell(row=row, column=col, value=r.controlling_lc).font = self.NORMAL_FONT; col += 1
            ws.cell(row=row, column=col, value=round(r.span_mm, 1)).font = self.NORMAL_FONT; col += 1

            if use_rel:
                ws.cell(row=row, column=col, value=r.ratio_str).font = self.NORMAL_FONT; col += 1
                c_d = ws.cell(row=row, column=col, value=round(r.max_deflection_mm, 3))
                c_d.font = self.NORMAL_FONT;  c_d.number_format = '0.000';  col += 1
                ws.cell(row=row, column=col, value=f"L/{int(r.allowable_ratio)}").font = self.NORMAL_FONT; col += 1
                rel_ok = r.rel_is_ok
                c_r = ws.cell(row=row, column=col, value="OK" if rel_ok else "NG")
                c_r.font = self.OK_FONT if rel_ok else self.NG_FONT
                c_r.fill = self.OK_FILL if rel_ok else self.NG_FILL
                col += 1

            if use_abs:
                c_abs = ws.cell(row=row, column=col, value=round(r.max_abs_deflection_mm, 3))
                c_abs.font = self.NORMAL_FONT;  c_abs.number_format = '0.000';  col += 1
                ws.cell(row=row, column=col, value=r.abs_limit_mm).font = self.NORMAL_FONT; col += 1
                abs_ok = r.abs_is_ok
                c_ar = ws.cell(row=row, column=col, value="OK" if abs_ok else "NG")
                c_ar.font = self.OK_FONT if abs_ok else self.NG_FONT
                c_ar.fill = self.OK_FILL if abs_ok else self.NG_FILL
                col += 1

            ws.cell(row=row, column=col, value=r.critical_node).font = self.NORMAL_FONT; col += 1
            el = ",".join(r.element_list) if r.element_list else ""
            ws.cell(row=row, column=col, value=f"({el})").font = self.NORMAL_FONT

            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).border = self.THIN_BORDER
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
            row += 1

        row += 1
        all_ok = all(_is_ok(r) for r in results)
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        ov = ws.cell(row=row, column=1,
                     value=f"OVERALL: {'ALL PASS ✓' if all_ok else 'SOME FAIL ✗'}")
        ov.font = Font(bold=True, size=12, name="Arial",
                       color="006100" if all_ok else "9C0006")
        ov.alignment = Alignment(horizontal="center")

        wb.save(filepath)
        return filepath


class TxtExporter:
    def export(self, results: List[BeamCheckSummary], output_dir: str,
               allowable_ratio: float = 360, abs_limit_mm: float = 25.0,
               use_rel: bool = True, use_abs: bool = True) -> str:
        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(output_dir, f"Deflection_Check_{ts}.txt")

        criteria_parts = []
        if use_rel:
            criteria_parts.append(f"Rel Limit: L/{int(allowable_ratio)}")
        if use_abs:
            criteria_parts.append(f"Abs Limit: {abs_limit_mm} mm")

        # Build header line
        hdr = f"  {'No':>6}  {'Section':>22}  {'Ctrl Combo':>12}  {'Span(mm)':>10}"
        if use_rel:
            hdr += f"  {'Defl':>10}  {'Criteria':>8}  {'Rel':>4}"
        if use_abs:
            hdr += f"  {'AbsMax(mm)':>10}  {'Lmt':>6}  {'Abs':>4}"
        hdr += f"  {'Ctrl Node':>12}  Elements"

        sep_len = max(110, len(hdr) + 4)
        lines = []
        lines.append("=" * sep_len)
        lines.append("                              Deflection Check Summary")
        lines.append("=" * sep_len)
        lines.append("  " + "   |   ".join(criteria_parts))
        lines.append(" ")
        lines.append(hdr)
        lines.append("=" * sep_len)

        def _sort_key(x):
            n = x.group_name.strip()
            return (0, int(n)) if n.isdigit() else (1, n)

        def _is_ok(r):
            if use_rel and use_abs:
                return r.rel_is_ok and r.abs_is_ok
            if use_rel:
                return r.rel_is_ok
            return r.abs_is_ok

        for r in sorted(results, key=_sort_key):
            el = f"({','.join(r.element_list)})" if r.element_list else ""
            row = (f"  {r.group_name:>6}"
                   f"  {r.section:>22}"
                   f"  {r.controlling_lc:>12}"
                   f"  {r.span_mm:>10.0f}")
            if use_rel:
                rel_str = "OK" if r.rel_is_ok else "NG"
                row += (f"  {r.ratio_str:>10}"
                        f"  L/{int(r.allowable_ratio):<6}"
                        f"  {rel_str:>4}")
            if use_abs:
                abs_str = "OK" if r.abs_is_ok else "NG"
                row += (f"  {r.max_abs_deflection_mm:>10.3f}"
                        f"  {r.abs_limit_mm:>6.1f}"
                        f"  {abs_str:>4}")
            row += f"  {r.critical_node:>12}  {el}"
            lines.append(row)

        all_ok = all(_is_ok(r) for r in results)
        lines.append("")
        lines.append(f"  Overall: {'ALL PASS' if all_ok else 'SOME FAIL'}")

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        return filepath
